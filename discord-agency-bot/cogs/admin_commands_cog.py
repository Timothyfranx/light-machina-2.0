import discord
from discord.ext import commands
from discord import app_commands
from pathlib import Path
import openpyxl
import shutil
import json
from datetime import datetime

from utils.storage_utils import Storage
from utils.excel_utils import get_user_excel_path

with open("config.json", "r") as f:
    CFG = json.load(f)

ADMIN_LOG_CHANNEL = int(CFG.get("ADMIN_CHANNEL_ID"))
GUILD_ID = int(CFG.get("GUILD_ID"))   # 👈 add this so we can bind commands to one server
REPORTS_DIR = Path("data/reports")
ARCHIVE_DIR = Path("data/archive")
ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)

storage = Storage("data/users.json")


class AdminCommandsCog(commands.Cog):
    def __init__(self, bot: commands.Bot):
        self.bot = bot

    # -------------------------------
    # 🔹 Delete user command
    # -------------------------------
    @app_commands.command(
        name="deleteuser",
        description="Admin: remove a user and archive their report"
    )
    @app_commands.guilds(discord.Object(id=GUILD_ID))  # 👈 force single-server registration
    @app_commands.checks.has_permissions(administrator=True)
    async def deleteuser(self, interaction: discord.Interaction, member: discord.Member):
        data = storage.get_user(str(member.id))
        if not data:
            return await interaction.response.send_message(
                "⚠️ User not tracked.", ephemeral=True
            )

        ch_id = data["channel_id"]
        username = data["username"]

        # Archive Excel
        p = get_user_excel_path(username)
        if p and p.exists():
            stamp = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
            dest = ARCHIVE_DIR / f"{username}-{stamp}.xlsx"
            shutil.copy2(p, dest)

        # Delete channel
        ch = interaction.guild.get_channel(int(ch_id))
        if ch:
            await ch.delete(reason="Admin removed user")

        # Remove from storage
        storage.remove_user(str(member.id))

        await interaction.response.send_message(
            f"🗑️ Removed {member.mention}, archived Excel.",
            ephemeral=True
        )

    # -------------------------------
    # 🔹 Compile all reports
    # -------------------------------
    @app_commands.command(
        name="getall",
        description="Admin: compile all reports into one master Excel"
    )
    @app_commands.guilds(discord.Object(id=GUILD_ID))  # 👈 force single-server registration
    @app_commands.checks.has_permissions(administrator=True)
    async def getall(self, interaction: discord.Interaction):
        await interaction.response.defer(ephemeral=True)

        master = openpyxl.Workbook()
        master.remove(master.active)

        for file in REPORTS_DIR.glob("*.xlsx"):
            try:
                wb = openpyxl.load_workbook(file, read_only=False)
                ws = wb.active
                new_ws = master.create_sheet(title=file.stem[:30])
                for row in ws.iter_rows(values_only=True):
                    new_ws.append(row)
            except Exception as e:
                print(f"⚠️ Skipping {file}: {e}")
                continue

        master_path = REPORTS_DIR / "master_report.xlsx"
        master.save(master_path)

        await interaction.followup.send(
            file=discord.File(master_path),
            ephemeral=True
        )


async def setup(bot: commands.Bot):
    await bot.add_cog(AdminCommandsCog(bot))
