"""
Admin Dashboard Cog

Provides an admin-only /dashboard command that shows:
- Total tracked users
- Total replies recorded across all users
- Average replies per user
- Top 5 users by replies
"""

import discord
from discord.ext import commands
from discord import app_commands
from datetime import datetime
import openpyxl
from pathlib import Path
import json

from utils.storage_utils import Storage
from utils.excel_utils import get_user_excel_path

# storage instance (same users.json used across cogs)
storage = Storage("data/users.json")

# load config robustly
with open("config.json", "r") as f:
    _CFG = json.load(f)
ADMIN_LOG_CHANNEL = (_CFG.get("ADMIN_CHANNEL_ID")
                     or _CFG.get("admin_log_channel")
                     or _CFG.get("admin_log_channel_id"))
REPORTS_DIR = Path("data/reports")


class AdminDashboardCog(commands.Cog):

    def __init__(self, bot: commands.Bot):
        self.bot = bot

    # -------------------------------------
    # Dashboard Command
    # -------------------------------------
    @app_commands.command(
        name="dashboard",
        description="Admin: show quick stats (total users, replies, top 5)")
    @app_commands.checks.has_permissions(administrator=True)
    async def dashboard(self, interaction: discord.Interaction):
        await interaction.response.defer(ephemeral=True)

        users = storage.list_users() or []  # defensive fallback

        total_users = len(users)
        total_replies = 0
        counts = {}

        for entry in users:
            try:
                # unpack tuple safely
                user_id, channel_id, username, target, start_date, status = entry
            except Exception:
                continue

            if not username:
                counts[username or str(user_id)] = 0
                continue

            p = get_user_excel_path(username)
            if not p or not p.exists():
                counts[username] = 0
                continue

            try:
                with openpyxl.load_workbook(p, read_only=True,
                                            data_only=True) as wb:
                    ws = wb.active
                    c = 0
                    for col in range(2, ws.max_column + 1):
                        for row in range(2, ws.max_row + 1):
                            if ws.cell(row=row,
                                       column=col).value not in (None, ""):
                                c += 1
                    counts[username] = c
                    total_replies += c
            except Exception as e:
                counts[username] = 0
                print(f"⚠️ Error reading {p}: {e}")

        avg_replies = round(total_replies /
                            total_users, 2) if total_users else 0.0
        top5 = sorted(counts.items(), key=lambda x: x[1], reverse=True)[:5]
        top_lines = ("\n".join(
            [f"**{name}** — {count}"
             for name, count in top5]) if top5 else "No data yet")

        embed = discord.Embed(title="📊 Agency Dashboard",
                              color=discord.Color.blurple())
        embed.add_field(name="Total Tracked Users",
                        value=str(total_users),
                        inline=True)
        embed.add_field(name="Total Replies Recorded",
                        value=str(total_replies),
                        inline=True)
        embed.add_field(name="Avg Replies / User",
                        value=str(avg_replies),
                        inline=True)
        embed.add_field(name="Top 5 Users", value=top_lines, inline=False)
        embed.set_footer(text=f"Generated {datetime.utcnow().isoformat()} UTC")

        await interaction.followup.send(embed=embed, ephemeral=True)

        # Admin log channel notification
        await self._log_event(
            f"📊 Dashboard viewed by {interaction.user.mention}")

    # -------------------------------------
    # Internal helper: log event
    # -------------------------------------
    async def _log_event(self, message: str):
        try:
            if ADMIN_LOG_CHANNEL:
                ch = self.bot.get_channel(int(ADMIN_LOG_CHANNEL))
                if ch:
                    await ch.send(message)
        except Exception as e:
            print(f"⚠️ Failed to log dashboard event: {e}")


async def setup(bot: commands.Bot):
    await bot.add_cog(AdminDashboardCog(bot))
