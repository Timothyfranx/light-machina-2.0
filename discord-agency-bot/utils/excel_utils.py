from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime, date, timedelta
import re
from typing import List, Optional

# Ensure reports directory exists
REPORTS_DIR = Path("data/reports")
REPORTS_DIR.mkdir(parents=True, exist_ok=True)


def _sanitize_filename(name: str) -> str:
    """Remove characters that break filenames and replace spaces with underscores."""
    return re.sub(r"[^\w\-_.()]", "", name).replace(" ", "_")


def create_user_excel(user_id: str | None, username: str, start_date: date,
                      end_date: date, target: int) -> Path:
    """
    Create a workbook for the user with date columns from start_date -> end_date inclusive.
    File name: REPORTS_DIR/{safe_username}.xlsx
    """
    if isinstance(start_date, str):
        start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
    if isinstance(end_date, str):
        end_date = datetime.strptime(end_date, "%Y-%m-%d").date()

    safe_username = _sanitize_filename(username)
    path = REPORTS_DIR / f"{safe_username}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Replies"

    # Header row: column A = "Day", then date columns
    ws.cell(row=1, column=1, value="Day")
    day = start_date
    col = 2
    while day <= end_date:
        ws.cell(row=1, column=col, value=day.isoformat())
        col += 1
        day += timedelta(days=1)

    # Metadata row for target
    ws.cell(row=2, column=1, value="Target")
    ws.cell(row=2, column=2, value=target)

    ws.row_dimensions[1].height = 20
    ws.sheet_view.showGridLines = True

    wb.save(path)
    return path


def get_user_excel_path(username: str) -> Optional[Path]:
    """Return path to the user’s Excel file if it exists."""
    safe_username = _sanitize_filename(username)
    path = REPORTS_DIR / f"{safe_username}.xlsx"
    return path if path.exists() else None


def _find_date_column(ws, date_iso: str) -> Optional[int]:
    """Find the column index for the given date string (YYYY-MM-DD)."""
    for col in range(2, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val == date_iso:
            return col
    return None


def record_links(username: str, target_date: date | datetime,
                 links: List[str]) -> bool:
    """
    Record links for the given username on target_date.
    Each link is stored as an Excel HYPERLINK with incremental numbering (1, 2, …).
    """
    if isinstance(target_date, datetime):
        target_date = target_date.date()
    date_iso = target_date.isoformat()

    safe_username = _sanitize_filename(username)
    path = get_user_excel_path(safe_username)
    if not path:
        raise FileNotFoundError(
            f"Excel for user '{username}' not found: expected {REPORTS_DIR}/{safe_username}.xlsx"
        )

    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Ensure column for this date exists
    col_idx = _find_date_column(ws, date_iso)
    if not col_idx:
        col_idx = ws.max_column + 1
        ws.cell(row=1, column=col_idx, value=date_iso)

    # Find next empty row (from row 3 down)
    start_row = 3
    row = start_row
    while ws.cell(row=row, column=col_idx).value not in (None, ""):
        row += 1

    existing_links = row - start_row
    idx = existing_links + 1

    # Write links
    for link in links:
        cell = ws.cell(row=row, column=col_idx)
        cell.value = f'=HYPERLINK("{link}", "{idx}")'
        cell.alignment = Alignment(horizontal="center")
        row += 1
        idx += 1

    wb.save(path)
    return True
